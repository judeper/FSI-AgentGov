"""
Excel Template Verification Script for FSI-AgentGov v1.6.2

Verifies all Excel files in docs/downloads/ for:
1. Correct file format (OOXML .xlsx, not DRM-encrypted OLE2)
2. Correct control counts per template
3. Stale version references (v1.0, v1.3.x, v1.4.0, April 2026)
4. Outdated control counts (48 controls)
5. Legacy path references (reference/pillar)
6. Dashboard Summary formulas present

Usage:
    python scripts/verify_excel_templates.py
"""

import os
import re
import sys
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

EXPECTED_COUNTS = {
    "governance-maturity-dashboard.xlsx": 78,
    "entra-administrator-checklist.xlsx": 5,
    "power-platform-administrator-checklist.xlsx": 8,
    "purview-administrator-checklist.xlsx": 7,
    "sharepoint-administrator-checklist.xlsx": 9,
    "compliance-officer-checklist.xlsx": 12,
}

STALE_PATTERNS = {
    "v1.0": "Outdated version reference",
    "v1.3.": "Outdated v1.3.x version reference",
    "v1.4.0": "Outdated version reference",
    "48 control": "Outdated control count",
    "April 2026": "Outdated release month",
    "reference/pillar": "Legacy path reference",
}

DASHBOARD_WORKBOOK = "governance-maturity-dashboard.xlsx"
DASHBOARD_SUMMARY_SHEET = "Summary"
CONTROL_INDEX_PATH = Path(__file__).resolve().parent.parent / "docs" / "controls" / "CONTROL-INDEX.md"
CONTROL_INDEX_PATTERN = re.compile(r"\|\s*(\d+\.\d+)\s*\|\s*\[([^\]]+)\]")


def load_canonical_control_titles(control_index_path):
    """Load canonical control titles from CONTROL-INDEX.md."""

    if not control_index_path.exists():
        return {}

    text = control_index_path.read_text(encoding="utf-8")
    return {
        match.group(1): match.group(2).strip()
        for match in CONTROL_INDEX_PATTERN.finditer(text)
    }


CANONICAL_CONTROL_TITLES = load_canonical_control_titles(CONTROL_INDEX_PATH)


def is_control_id(value):
    """Return True when a cell value looks like a framework control ID."""

    cell_value = str(value).strip()
    if "." not in cell_value:
        return False

    parts = cell_value.split(".")
    return len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit()


def count_formulas(sheet):
    """Count formula cells on a worksheet."""

    formulas = 0
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formulas += 1
    return formulas


def verify_excel_file(file_path):
    """Verify a single Excel file for format, control counts and stale content."""

    filename = os.path.basename(file_path)
    print(f"\n{'=' * 80}")
    print(f"Verifying: {filename}")
    print(f"{'=' * 80}")

    issues = []

    if not zipfile.is_zipfile(file_path):
        issues.append(
            "[FAIL] File is not valid OOXML (.xlsx) format — appears to be DRM-encrypted (OLE2 container). "
            "Remove the sensitivity label in Excel and re-save as .xlsx to fix."
        )
        print("[FAIL] DRM-encrypted OLE2 format detected — not readable as .xlsx")
        print("   Fix: Open in Excel → File → Info → Remove sensitivity label → Save As .xlsx")
        return issues

    try:
        workbook = load_workbook(file_path, data_only=False)

        total_controls = 0
        control_breakdown = {}

        for sheet in workbook.worksheets:
            sheet_control_count = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row and row[0] and is_control_id(row[0]):
                    sheet_control_count += 1
                    total_controls += 1

            if sheet_control_count > 0:
                control_breakdown[sheet.title] = sheet_control_count

        expected_count = EXPECTED_COUNTS.get(filename)
        if expected_count is None:
            print(f"[WARN] Control count: {total_controls} (no expected count defined)")
        elif total_controls != expected_count:
            issues.append(
                f"[FAIL] Control count mismatch: Found {total_controls}, expected {expected_count}"
            )
            print(f"[FAIL] Control count: {total_controls} (expected {expected_count})")
        else:
            print(f"[PASS] Control count: {total_controls} (correct)")

        if control_breakdown:
            print("\n   Control breakdown by sheet:")
            for sheet_name, count in control_breakdown.items():
                print(f"   - {sheet_name}: {count} controls")

        print("\n   Checking control titles...")
        title_mismatches = []
        for sheet in workbook.worksheets:
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if not row or len(row) < 2:
                    continue

                control_id = row[0]
                control_name = row[1]
                if not control_id or not is_control_id(control_id) or not isinstance(control_name, str):
                    continue

                canonical_title = CANONICAL_CONTROL_TITLES.get(str(control_id).strip())
                if canonical_title and control_name.strip() != canonical_title:
                    title_mismatches.append(
                        {
                            "sheet": sheet.title,
                            "cell": f"B{row_idx}",
                            "control_id": str(control_id).strip(),
                            "value": control_name[:100],
                            "expected": canonical_title[:100],
                        }
                    )

        if title_mismatches:
            issues.append(
                f"[FAIL] Found {len(title_mismatches)} stale control title(s) that do not match CONTROL-INDEX.md"
            )
            for mismatch in title_mismatches[:5]:
                print(
                    "   [FAIL] "
                    f"{mismatch['sheet']}!{mismatch['cell']} {mismatch['control_id']}: "
                    f"'{mismatch['value']}' != '{mismatch['expected']}'"
                )
            if len(title_mismatches) > 5:
                print(f"   ... and {len(title_mismatches) - 5} more")
        else:
            print("   [PASS] Control titles match CONTROL-INDEX.md")

        print("\n   Checking for stale content...")
        stale_findings = {pattern: [] for pattern in STALE_PATTERNS}

        for sheet in workbook.worksheets:
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                for col_idx, cell_value in enumerate(row, start=1):
                    if cell_value is None:
                        continue

                    cell_str = str(cell_value).lower()
                    for pattern in STALE_PATTERNS:
                        if pattern.lower() in cell_str:
                            stale_findings[pattern].append(
                                {
                                    "sheet": sheet.title,
                                    "cell": f"{get_column_letter(col_idx)}{row_idx}",
                                    "value": str(cell_value)[:100],
                                }
                            )

        stale_found = False
        for pattern, findings in stale_findings.items():
            if not findings:
                continue

            stale_found = True
            description = STALE_PATTERNS[pattern]
            issues.append(f"[FAIL] Found '{pattern}' ({description}) in {len(findings)} location(s)")
            print(f"\n   [FAIL] Found '{pattern}' ({description}):")
            for finding in findings[:5]:
                print(f"      - {finding['sheet']}!{finding['cell']}: {finding['value']}")
            if len(findings) > 5:
                print(f"      ... and {len(findings) - 5} more")

        if not stale_found:
            print("   [PASS] No stale content patterns found")

        if filename == DASHBOARD_WORKBOOK:
            summary_sheet = workbook[DASHBOARD_SUMMARY_SHEET]
            summary_formula_count = count_formulas(summary_sheet)
            if summary_formula_count == 0:
                issues.append("[FAIL] Dashboard Summary sheet has 0 formulas")
                print("\n   [FAIL] Dashboard Summary sheet has 0 formulas")
            else:
                print(
                    f"\n   [PASS] Dashboard Summary sheet formula count: {summary_formula_count}"
                )

        workbook.close()

    except Exception as exc:
        issues.append(f"[FAIL] Error reading file: {exc}")
        print(f"[FAIL] Error: {exc}")

    return issues


def main():
    """Main verification function."""

    script_dir = Path(__file__).parent
    project_root = script_dir.parent
    downloads_dir = project_root / "docs" / "downloads"

    print("FSI-AgentGov Excel Template Verification")
    print(f"{'=' * 80}")
    print(f"Project root: {project_root}")
    print(f"Downloads dir: {downloads_dir}")

    if not downloads_dir.exists():
        print(f"\n[FAIL] Downloads directory not found: {downloads_dir}")
        return 1

    excel_files = list(downloads_dir.glob("*.xlsx"))
    if not excel_files:
        print(f"\n[FAIL] No Excel files found in {downloads_dir}")
        return 1

    print(f"\nFound {len(excel_files)} Excel file(s) to verify")

    all_issues = {}
    for excel_file in sorted(excel_files):
        issues = verify_excel_file(excel_file)
        if issues:
            all_issues[excel_file.name] = issues

    print(f"\n{'=' * 80}")
    print("VERIFICATION SUMMARY")
    print(f"{'=' * 80}")

    if all_issues:
        print(f"\n[FAIL] Issues found in {len(all_issues)} file(s):\n")
        for filename, issues in all_issues.items():
            print(f"{filename}:")
            for issue in issues:
                print(f"  {issue}")
            print()
        return 1

    print("\n[PASS] All Excel files passed verification!")
    print("   - Control counts correct")
    print("   - No stale content found")
    print("   - Dashboard formulas present")
    return 0


if __name__ == "__main__":
    sys.exit(main())
