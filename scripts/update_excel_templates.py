"""
Excel Template Update Script for FSI-AgentGov v1.1

Updates all Excel files in docs/downloads/:
1. Changes "v1.0 Beta" to "v1.1" in all files
2. Identifies missing controls in governance-maturity-dashboard.xlsx
3. Adds missing controls to the dashboard

Usage:
    python scripts/update_excel_templates.py --check    # Preview changes only
    python scripts/update_excel_templates.py --update   # Apply changes

Author: Claude Code
"""

import os
import sys
import argparse
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# All 61 controls in the framework
ALL_CONTROLS = {
    # Pillar 1 - Security (23 controls)
    "1.1": "Restrict Agent Publishing by Authorization",
    "1.2": "Agent Registry and Integrated Apps Management",
    "1.3": "SharePoint Content Governance and Permissions",
    "1.4": "Advanced Connector Policies (ACP)",
    "1.5": "Data Loss Prevention (DLP) and Sensitivity Labels",
    "1.6": "Microsoft Purview DSPM for AI",
    "1.7": "Comprehensive Audit Logging and Compliance",
    "1.8": "Runtime Protection and External Threat Detection",
    "1.9": "Data Retention and Deletion Policies",
    "1.10": "Communication Compliance Monitoring",
    "1.11": "Conditional Access and Phishing-Resistant MFA",
    "1.12": "Insider Risk Detection and Response",
    "1.13": "Sensitive Information Types (SITs) and Pattern Recognition",
    "1.14": "Data Minimization and Agent Scope Control",
    "1.15": "Encryption (Data in Transit and at Rest)",
    "1.16": "Information Rights Management (IRM) for Documents",
    "1.17": "Endpoint Data Loss Prevention (Endpoint DLP)",
    "1.18": "Application-Level Authorization and Role-Based Access Control (RBAC)",
    "1.19": "eDiscovery for Agent Interactions",
    "1.20": "Network Isolation / Private Connectivity",
    "1.21": "Adversarial Input Logging",
    "1.22": "Information Barriers",
    "1.23": "Step-Up Authentication for Agent Operations",
    # Pillar 2 - Management (21 controls)
    "2.1": "Managed Environments",
    "2.2": "Environment Groups and Tier Classification",
    "2.3": "Change Management and Release Planning",
    "2.4": "Business Continuity and Disaster Recovery",
    "2.5": "Testing, Validation, and Quality Assurance",
    "2.6": "Model Risk Management Alignment with OCC 2011-12 / SR 11-7",
    "2.7": "Vendor and Third-Party Risk Management",
    "2.8": "Access Control and Segregation of Duties",
    "2.9": "Agent Performance Monitoring and Optimization",
    "2.10": "Patch Management and System Updates",
    "2.11": "Bias Testing and Fairness Assessment (FINRA Notice 25-07 / SR 11-7 Alignment)",
    "2.12": "Supervision and Oversight (FINRA Rule 3110)",
    "2.13": "Documentation and Record Keeping",
    "2.14": "Training and Awareness Program",
    "2.15": "Environment Routing",
    "2.16": "RAG Source Integrity Validation",
    "2.17": "Multi-Agent Orchestration Limits",
    "2.18": "Automated Conflict of Interest Testing",
    "2.19": "Customer AI Disclosure and Transparency",
    "2.20": "Adversarial Testing and Red Team Framework",
    "2.21": "AI Marketing Claims and Substantiation",
    # Pillar 3 - Reporting (10 controls)
    "3.1": "Agent Inventory and Metadata Management",
    "3.2": "Usage Analytics and Activity Monitoring",
    "3.3": "Compliance and Regulatory Reporting",
    "3.4": "Incident Reporting and Root Cause Analysis",
    "3.5": "Cost Allocation and Budget Tracking",
    "3.6": "Orphaned Agent Detection and Remediation",
    "3.7": "PPAC Security Posture Assessment",
    "3.8": "Copilot Hub and Governance Dashboard",
    "3.9": "Microsoft Sentinel Integration",
    "3.10": "Hallucination Feedback Loop",
    # Pillar 4 - SharePoint (7 controls)
    "4.1": "SharePoint Information Access Governance (IAG) / Restricted Content Discovery",
    "4.2": "Site Access Reviews and Certification",
    "4.3": "Site and Document Retention Management",
    "4.4": "Guest and External User Access Controls",
    "4.5": "SharePoint Security and Compliance Monitoring",
    "4.6": "Grounding Scope Governance",
    "4.7": "Microsoft 365 Copilot Data Governance",
}

PILLAR_NAMES = {
    "1": "Pillar 1 - Security",
    "2": "Pillar 2 - Management",
    "3": "Pillar 3 - Reporting",
    "4": "Pillar 4 - SharePoint"
}


def get_controls_in_file(file_path):
    """Extract all control IDs from an Excel file."""
    controls_found = set()

    wb = load_workbook(file_path, data_only=True)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            if row and row[0]:
                cell_value = str(row[0]).strip()
                # Match pattern like "1.1", "2.15", "3.10", "4.7"
                if '.' in cell_value and len(cell_value.split('.')) == 2:
                    parts = cell_value.split('.')
                    if parts[0].isdigit() and parts[1].isdigit():
                        controls_found.add(cell_value)

    wb.close()
    return controls_found


def find_version_cells(file_path):
    """Find cells containing version references."""
    version_cells = []

    wb = load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row_idx, row in enumerate(sheet.iter_rows(), start=1):
            for col_idx, cell in enumerate(row, start=1):
                if cell.value and "v1.0" in str(cell.value).lower():
                    version_cells.append({
                        "sheet": sheet_name,
                        "row": row_idx,
                        "col": col_idx,
                        "current_value": str(cell.value),
                        "cell_ref": f"{sheet_name}!{cell.coordinate}"
                    })

    wb.close()
    return version_cells


def update_version_references(file_path, dry_run=True):
    """Update all v1.0 references to v1.1."""
    version_cells = find_version_cells(file_path)

    if not version_cells:
        return 0, []

    if dry_run:
        return len(version_cells), version_cells

    # Actually update
    wb = load_workbook(file_path)

    for vc in version_cells:
        sheet = wb[vc["sheet"]]
        cell = sheet.cell(row=vc["row"], column=vc["col"])

        # Replace v1.0 Beta with v1.1
        new_value = str(cell.value).replace("v1.0 Beta", "v1.1").replace("v1.0", "v1.1")
        cell.value = new_value
        vc["new_value"] = new_value

    wb.save(file_path)
    wb.close()

    return len(version_cells), version_cells


def find_missing_controls(file_path):
    """Find controls missing from governance-maturity-dashboard.xlsx."""
    controls_found = get_controls_in_file(file_path)
    all_control_ids = set(ALL_CONTROLS.keys())

    missing = all_control_ids - controls_found
    extra = controls_found - all_control_ids

    return sorted(missing, key=lambda x: (int(x.split('.')[0]), int(x.split('.')[1]))), list(extra)


def add_missing_controls(file_path, missing_controls, dry_run=True):
    """Add missing controls to the All Controls sheet."""
    if not missing_controls:
        return []

    if dry_run:
        return missing_controls

    wb = load_workbook(file_path)

    # Find the "All Controls" sheet
    if "All Controls" not in wb.sheetnames:
        print(f"[ERROR] 'All Controls' sheet not found in {file_path}")
        wb.close()
        return []

    sheet = wb["All Controls"]

    # Find the last row with data in each pillar section
    # First, find all existing control rows and their positions
    pillar_sections = {}  # pillar -> list of (row_idx, control_id)

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row and row[0]:
            cell_value = str(row[0]).strip()
            if '.' in cell_value and len(cell_value.split('.')) == 2:
                parts = cell_value.split('.')
                if parts[0].isdigit() and parts[1].isdigit():
                    pillar = parts[0]
                    if pillar not in pillar_sections:
                        pillar_sections[pillar] = []
                    pillar_sections[pillar].append((row_idx, cell_value))

    # Group missing controls by pillar
    missing_by_pillar = {}
    for ctrl_id in missing_controls:
        pillar = ctrl_id.split('.')[0]
        if pillar not in missing_by_pillar:
            missing_by_pillar[pillar] = []
        missing_by_pillar[pillar].append(ctrl_id)

    added_controls = []

    # Sort pillars in reverse order so row insertions don't affect other pillar positions
    for pillar in sorted(missing_by_pillar.keys(), reverse=True):
        missing_ids = missing_by_pillar[pillar]
        if pillar not in pillar_sections:
            print(f"[WARN] Pillar {pillar} section not found in spreadsheet")
            continue

        existing_controls = pillar_sections[pillar]
        last_row = max(row for row, _ in existing_controls)

        # Sort missing controls by their number (e.g., 1.23 comes after 1.22)
        sorted_missing = sorted(missing_ids, key=lambda x: int(x.split('.')[1]))

        for i, ctrl_id in enumerate(sorted_missing):
            insert_row = last_row + 1 + i

            # Insert a new row to avoid merged cell issues
            sheet.insert_rows(insert_row)

            # Now write to the newly inserted row (which won't be merged)
            sheet.cell(row=insert_row, column=1, value=ctrl_id)
            sheet.cell(row=insert_row, column=2, value=ALL_CONTROLS.get(ctrl_id, ""))
            sheet.cell(row=insert_row, column=3, value="Not Started")  # Status
            sheet.cell(row=insert_row, column=4, value="")  # Notes
            sheet.cell(row=insert_row, column=5, value="")  # Due Date

            added_controls.append(ctrl_id)
            print(f"      Inserted control {ctrl_id} at row {insert_row}")

    wb.save(file_path)
    wb.close()

    return added_controls


def update_summary_count(file_path, dry_run=True):
    """Update the total control count in Summary Dashboard sheet."""
    wb = load_workbook(file_path)

    if "Summary Dashboard" not in wb.sheetnames:
        wb.close()
        return None

    sheet = wb["Summary Dashboard"]

    # Find cells containing "58" or total count
    count_cells = []
    for row_idx, row in enumerate(sheet.iter_rows(), start=1):
        for col_idx, cell in enumerate(row, start=1):
            if cell.value:
                val_str = str(cell.value)
                if val_str == "58" or "58 control" in val_str.lower():
                    count_cells.append({
                        "sheet": "Summary Dashboard",
                        "row": row_idx,
                        "col": col_idx,
                        "current_value": val_str,
                        "cell_ref": cell.coordinate
                    })

    if dry_run:
        wb.close()
        return count_cells

    # Update any found cells
    for cc in count_cells:
        cell = sheet.cell(row=cc["row"], column=cc["col"])
        new_value = str(cell.value).replace("58", "60")
        cell.value = new_value
        cc["new_value"] = new_value

    wb.save(file_path)
    wb.close()

    return count_cells


def main():
    parser = argparse.ArgumentParser(description="Update Excel templates for FSI-AgentGov v1.1")
    parser.add_argument("--check", action="store_true", help="Preview changes only (dry run)")
    parser.add_argument("--update", action="store_true", help="Apply changes")
    args = parser.parse_args()

    if not args.check and not args.update:
        print("Usage:")
        print("  python scripts/update_excel_templates.py --check   # Preview changes")
        print("  python scripts/update_excel_templates.py --update  # Apply changes")
        return 1

    dry_run = args.check

    # Get paths
    script_dir = Path(__file__).parent
    project_root = script_dir.parent
    downloads_dir = project_root / "docs" / "downloads"

    print(f"FSI-AgentGov Excel Template Update Script")
    print(f"{'='*80}")
    print(f"Mode: {'DRY RUN (preview only)' if dry_run else 'APPLYING CHANGES'}")
    print(f"Downloads dir: {downloads_dir}")

    # Find all Excel files
    excel_files = list(downloads_dir.glob("*.xlsx"))

    if not excel_files:
        print(f"\n[ERROR] No Excel files found")
        return 1

    print(f"\nFound {len(excel_files)} Excel file(s)")

    # Track all changes
    all_changes = {}

    # Process each file
    for excel_file in sorted(excel_files):
        filename = excel_file.name
        print(f"\n{'='*80}")
        print(f"Processing: {filename}")
        print(f"{'='*80}")

        file_changes = {"version_updates": [], "missing_controls": [], "added_controls": []}

        # 1. Check/update version references
        count, version_cells = update_version_references(excel_file, dry_run=dry_run)
        if count > 0:
            print(f"\n[VERSION] Found {count} version reference(s) to update:")
            for vc in version_cells:
                if dry_run:
                    print(f"   {vc['cell_ref']}: '{vc['current_value']}' -> v1.1")
                else:
                    print(f"   {vc['cell_ref']}: Updated to '{vc.get('new_value', 'v1.1')}'")
            file_changes["version_updates"] = version_cells
        else:
            print(f"\n[VERSION] No version references to update")

        # 2. For governance-maturity-dashboard.xlsx, check for missing controls
        if filename == "governance-maturity-dashboard.xlsx":
            missing, extra = find_missing_controls(excel_file)

            if missing:
                print(f"\n[CONTROLS] Missing {len(missing)} control(s):")
                for ctrl_id in missing:
                    print(f"   - {ctrl_id}: {ALL_CONTROLS.get(ctrl_id, 'Unknown')}")
                file_changes["missing_controls"] = missing

                if not dry_run:
                    added = add_missing_controls(excel_file, missing, dry_run=False)
                    if added:
                        print(f"\n[CONTROLS] Added {len(added)} control(s) to spreadsheet")
                        file_changes["added_controls"] = added
            else:
                print(f"\n[CONTROLS] All 61 controls present")

            if extra:
                print(f"\n[WARN] Found {len(extra)} unexpected control ID(s): {extra}")

            # Check summary count
            count_cells = update_summary_count(excel_file, dry_run=dry_run)
            if count_cells:
                print(f"\n[SUMMARY] Found count cells to update:")
                for cc in count_cells:
                    print(f"   {cc['cell_ref']}: '{cc['current_value']}'")

        if any(file_changes.values()):
            all_changes[filename] = file_changes

    # Summary
    print(f"\n{'='*80}")
    print(f"SUMMARY")
    print(f"{'='*80}")

    if all_changes:
        total_version = sum(len(c["version_updates"]) for c in all_changes.values())
        total_missing = sum(len(c["missing_controls"]) for c in all_changes.values())

        if dry_run:
            print(f"\nChanges to apply:")
            print(f"   - {total_version} version reference(s) to update")
            print(f"   - {total_missing} missing control(s) to add")
            print(f"\nRun with --update to apply these changes")
        else:
            print(f"\nChanges applied:")
            print(f"   - {total_version} version reference(s) updated")
            total_added = sum(len(c["added_controls"]) for c in all_changes.values())
            print(f"   - {total_added} control(s) added")
            print(f"\n[OK] All Excel files updated successfully!")
    else:
        print(f"\n[OK] No changes needed - all files are up to date")

    return 0


if __name__ == "__main__":
    sys.exit(main())
