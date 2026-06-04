"""
Generate FSI-AgentGov Excel checklist templates.
Creates 6 professional, clean .xlsx files with no DRM/encryption.
"""

import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── All 78 controls ──────────────────────────────────────────────────────────

ALL_CONTROLS = {
    "Pillar 1 \u2014 Security": [
        ("1.1", "Restrict Agent Publishing by Authorization"),
        ("1.2", "Agent Registry and Integrated Apps Management"),
        ("1.3", "SharePoint Content Governance and Permissions"),
        ("1.4", "Advanced Connector Policies (ACP)"),
        ("1.5", "Data Loss Prevention (DLP) and Sensitivity Labels"),
        ("1.6", "Microsoft Purview: DSPM for AI"),
        ("1.7", "Comprehensive Audit Logging and Compliance"),
        ("1.8", "Runtime Protection and External Threat Detection"),
        ("1.9", "Data Retention and Deletion Policies"),
        ("1.10", "Communication Compliance Monitoring"),
        ("1.11", "Conditional Access and Phishing-Resistant MFA"),
        ("1.12", "Insider Risk Detection and Response"),
        ("1.13", "Sensitive Information Types (SITs) and Pattern Recognition"),
        ("1.14", "Data Minimization and Agent Scope Control"),
        ("1.15", "Encryption: Data in Transit and at Rest"),
        ("1.16", "Information Rights Management (IRM) for Documents"),
        ("1.17", "Endpoint Data Loss Prevention (Endpoint DLP)"),
        ("1.18", "Application-Level Authorization and RBAC"),
        ("1.19", "eDiscovery for Agent Interactions"),
        ("1.20", "Network Isolation and Private Connectivity"),
        ("1.21", "Adversarial Input Logging"),
        ("1.22", "Information Barriers for AI Agents"),
        ("1.23", "Step-Up Authentication for AI Agent Operations"),
        ("1.24", "Defender AI Security Posture Management (AI-SPM)"),
        ("1.25", "MIME Type Restrictions for File Uploads"),
        ("1.26", "Agent File Upload and File Analysis Restrictions"),
        ("1.27", "AI Agent Content Moderation Enforcement"),
        ("1.28", "Policy-Based Agent Publishing Restrictions"),
        ("1.29", "Global Secure Access: Network Controls for Copilot Studio Agents"),
    ],
    "Pillar 2 \u2014 Management": [
        ("2.1", "Managed Environments"),
        ("2.2", "Environment Groups and Tier Classification"),
        ("2.3", "Change Management and Release Planning"),
        ("2.4", "Business Continuity and Disaster Recovery"),
        ("2.5", "Testing, Validation, and Quality Assurance"),
        ("2.6", "Model Risk Management (OCC Bulletin 2026-13 / Fed SR 26-2)"),
        ("2.7", "Vendor and Third-Party Risk Management"),
        ("2.8", "Access Control and Segregation of Duties"),
        ("2.9", "Agent Performance Monitoring and Optimization"),
        ("2.10", "Patch Management and System Updates"),
        ("2.11", "Bias Testing and Fairness Assessment"),
        ("2.12", "Supervision and Oversight (FINRA Rule 3110)"),
        ("2.13", "Documentation and Record Keeping"),
        ("2.14", "Training and Awareness Program"),
        ("2.15", "Environment Routing and Auto-Provisioning"),
        ("2.16", "RAG Source Integrity Validation"),
        ("2.17", "Multi-Agent Orchestration Limits"),
        ("2.18", "Automated Conflict of Interest Testing"),
        ("2.19", "Customer AI Disclosure and Transparency"),
        ("2.20", "Adversarial Testing and Red Team Framework"),
        ("2.21", "AI Marketing Claims and Substantiation"),
        ("2.22", "Inactivity Timeout Enforcement"),
        ("2.23", "User Consent and AI Disclosure Enforcement"),
        ("2.24", "Agent Feature Enablement and Restriction Governance"),
        ("2.25", "Microsoft Agent 365: Admin Center Governance Console"),
        ("2.26", "Entra Agent ID: Identity Governance for Agents"),
    ],
    "Pillar 3 \u2014 Reporting": [
        ("3.1", "Agent Inventory and Metadata Management"),
        ("3.2", "Usage Analytics and Activity Monitoring"),
        ("3.3", "Compliance and Regulatory Reporting"),
        ("3.4", "Incident Reporting and Root Cause Analysis"),
        ("3.5", "Cost Allocation and Budget Tracking"),
        ("3.6", "Orphaned Agent Detection and Remediation"),
        ("3.7", "PPAC Security Posture Assessment"),
        ("3.8", "Copilot Hub and Governance Dashboard"),
        ("3.9", "Microsoft Sentinel Integration"),
        ("3.10", "Hallucination Feedback Loop"),
        ("3.11", "Centralized Agent Inventory Enforcement"),
        ("3.12", "Agent Governance Exception and Override Management"),
        ("3.13", "Agent 365 Admin Center Analytics and Reporting"),
        ("3.14", "Agent 365 Observability SDK and Custom Agent Telemetry"),
    ],
    "Pillar 4 \u2014 SharePoint": [
        ("4.1", "SharePoint Information Access Governance (IAG)"),
        ("4.2", "Site Access Reviews and Certification"),
        ("4.3", "Site and Document Retention Management"),
        ("4.4", "Guest and External User Access Controls"),
        ("4.5", "SharePoint Security and Compliance Monitoring"),
        ("4.6", "Grounding Scope Governance"),
        ("4.7", "Microsoft 365 Copilot Data Governance"),
        ("4.8", "Item-Level Permission Scanning for Agent Knowledge Sources"),
        ("4.9", "Embedded File Content Governance"),
    ],
}

# Build flat lookup
CONTROL_LOOKUP = {}
for _pillar, controls in ALL_CONTROLS.items():
    for cid, cname in controls:
        CONTROL_LOOKUP[cid] = cname

# ── Role-specific templates ──────────────────────────────────────────────────

ROLE_TEMPLATES = {
    "entra-administrator-checklist.xlsx": {
        "title": "Entra Administrator Checklist",
        "role": "Entra Global Admin",
        "controls": ["1.11", "1.12", "1.18", "2.26", "3.1"],
    },
    "power-platform-administrator-checklist.xlsx": {
        "title": "Power Platform Administrator Checklist",
        "role": "Power Platform Admin",
        "controls": ["1.29", "2.1", "2.2", "2.15", "2.16", "2.17", "3.7", "3.8"],
    },
    "purview-administrator-checklist.xlsx": {
        "title": "Purview Administrator Checklist",
        "role": "Purview Compliance Admin",
        "controls": ["1.5", "1.6", "1.7", "1.9", "1.10", "1.19", "1.22"],
    },
    "sharepoint-administrator-checklist.xlsx": {
        "title": "SharePoint Administrator Checklist",
        "role": "SharePoint Admin",
        "controls": ["4.1", "4.2", "4.3", "4.4", "4.5", "4.6", "4.7", "4.8", "4.9"],
    },
    "compliance-officer-checklist.xlsx": {
        "title": "Compliance Officer Checklist",
        "role": "Compliance Officer",
        "controls": [
            "1.7", "1.19", "1.22", "2.6", "2.11", "2.12",
            "2.13", "2.18", "2.19", "2.21", "3.3", "3.10",
        ],
    },
}

# ── Colour palette ───────────────────────────────────────────────────────────

PRIMARY_DARK = "1F4E79"     # Dark blue for headers
PRIMARY_MED  = "2E75B6"     # Medium blue for accents
PRIMARY_LIGHT = "D6E4F0"    # Light blue for alternating rows
ACCENT_GREEN = "E2EFDA"     # Light green for status column
WHITE = "FFFFFF"
LIGHT_GREY = "F2F2F2"
MED_GREY = "808080"

# ── Reusable styles ──────────────────────────────────────────────────────────

TITLE_FONT = Font(name="Aptos", bold=True, size=16, color=PRIMARY_DARK)
SUBTITLE_FONT = Font(name="Aptos", size=10, color=MED_GREY)
HEADER_FONT = Font(name="Aptos", bold=True, size=11, color=WHITE)
HEADER_FILL = PatternFill(start_color=PRIMARY_DARK, end_color=PRIMARY_DARK, fill_type="solid")
DATA_FONT = Font(name="Aptos", size=11)
DATA_FONT_BOLD = Font(name="Aptos", size=11, bold=True, color=PRIMARY_DARK)
FOOTER_FONT = Font(name="Aptos", italic=True, size=9, color=MED_GREY)
EVEN_ROW_FILL = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")
PILLAR_FILL = PatternFill(start_color=PRIMARY_LIGHT, end_color=PRIMARY_LIGHT, fill_type="solid")
PILLAR_FONT = Font(name="Aptos", bold=True, size=11, color=PRIMARY_DARK)

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
HEADER_BORDER = Border(
    left=Side(style="thin", color=PRIMARY_DARK),
    right=Side(style="thin", color=PRIMARY_DARK),
    top=Side(style="thin", color=PRIMARY_DARK),
    bottom=Side(style="medium", color=PRIMARY_DARK),
)

HEADERS = ["Control ID", "Control Name", "Status", "Notes", "Due Date"]
COL_WIDTHS = [14, 56, 18, 45, 16]
STATUS_OPTIONS = '"Not Started,In Progress,Completed,N/A"'
VERSION_FOOTER = "FSI Agent Governance Framework v1.3.0 \u2014 March 2026"


def style_header_row(ws, row):
    """Apply professional header styling."""
    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS, strict=False), 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = HEADER_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[row].height = 24


def write_data_row(ws, row, cid, cname, is_even):
    """Write a single control row with alternating shading."""
    fill = EVEN_ROW_FILL if is_even else PatternFill()
    
    # Control ID — bold, centered
    cell_id = ws.cell(row=row, column=1, value=cid)
    cell_id.font = DATA_FONT_BOLD
    cell_id.alignment = Alignment(horizontal="center", vertical="center")
    cell_id.border = THIN_BORDER
    cell_id.fill = fill

    # Control Name
    cell_name = ws.cell(row=row, column=2, value=cname)
    cell_name.font = DATA_FONT
    cell_name.alignment = Alignment(vertical="center", wrap_text=True)
    cell_name.border = THIN_BORDER
    cell_name.fill = fill

    # Status — default "Not Started"
    cell_status = ws.cell(row=row, column=3, value="Not Started")
    cell_status.font = DATA_FONT
    cell_status.alignment = Alignment(horizontal="center", vertical="center")
    cell_status.border = THIN_BORDER
    cell_status.fill = fill

    # Notes
    cell_notes = ws.cell(row=row, column=4)
    cell_notes.font = DATA_FONT
    cell_notes.alignment = Alignment(vertical="center", wrap_text=True)
    cell_notes.border = THIN_BORDER
    cell_notes.fill = fill

    # Due Date
    cell_date = ws.cell(row=row, column=5)
    cell_date.font = DATA_FONT
    cell_date.alignment = Alignment(horizontal="center", vertical="center")
    cell_date.number_format = "YYYY-MM-DD"
    cell_date.border = THIN_BORDER
    cell_date.fill = fill

    ws.row_dimensions[row].height = 22


def add_status_validation(ws, start_row, end_row):
    """Add dropdown validation for Status column."""
    dv = DataValidation(
        type="list",
        formula1=STATUS_OPTIONS,
        allow_blank=True,
        showErrorMessage=True,
    )
    dv.error = "Please select: Not Started, In Progress, Completed, or N/A"
    dv.errorTitle = "Invalid Status"
    dv.prompt = "Select implementation status"
    dv.promptTitle = "Status"
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(f"C{start_row}:C{end_row}")


def add_footer(ws, row):
    """Add version footer."""
    ws.merge_cells(f"A{row}:E{row}")
    cell = ws[f"A{row}"]
    cell.value = VERSION_FOOTER
    cell.font = FOOTER_FONT
    cell.alignment = Alignment(horizontal="center")


def freeze_and_filter(ws, header_row):
    """Freeze panes below header and enable auto-filter."""
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:E{header_row}"


def set_print_settings(ws, title):
    """Configure print layout."""
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape"
    ws.oddHeader.center.text = title
    ws.oddFooter.center.text = VERSION_FOOTER


# ── Build role-specific checklists ───────────────────────────────────────────

def create_role_checklist(filename, config):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = config["role"]

    # Title block
    ws.merge_cells("A1:E1")
    ws["A1"].value = config["title"]
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:E2")
    role_text = "Role: " + config["role"] + "  |  " + VERSION_FOOTER
    ws["A2"].value = role_text
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22

    # Blank spacer row
    ws.row_dimensions[3].height = 6

    # Header row
    header_row = 4
    style_header_row(ws, header_row)

    # Data rows
    data_start = header_row + 1
    row = data_start
    for i, cid in enumerate(config["controls"]):
        cname = CONTROL_LOOKUP.get(cid, "Unknown")
        write_data_row(ws, row, cid, cname, is_even=(i % 2 == 0))
        row += 1

    add_status_validation(ws, data_start, row - 1)
    freeze_and_filter(ws, header_row)
    set_print_settings(ws, config["title"])

    # Footer
    add_footer(ws, row + 1)

    path = os.path.join("docs", "downloads", filename)
    wb.save(path)
    print(f"  Created: {filename} ({len(config['controls'])} controls)")


# ── Build governance maturity dashboard ──────────────────────────────────────

def create_dashboard():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Summary sheet
    ws_summary = wb.create_sheet(title="Summary")
    ws_summary.merge_cells("A1:E1")
    ws_summary["A1"].value = "Governance Maturity Dashboard"
    ws_summary["A1"].font = TITLE_FONT
    ws_summary["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws_summary.row_dimensions[1].height = 36

    ws_summary.merge_cells("A2:E2")
    ws_summary["A2"].value = VERSION_FOOTER
    ws_summary["A2"].font = SUBTITLE_FONT
    ws_summary.row_dimensions[2].height = 22

    # Summary table
    ws_summary.row_dimensions[4].height = 6
    summary_headers = ["Pillar", "Controls", "Not Started", "In Progress", "Completed"]
    summary_widths = [30, 14, 14, 14, 14]
    for col_idx, (h, w) in enumerate(zip(summary_headers, summary_widths, strict=False), 1):
        cell = ws_summary.cell(row=5, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = HEADER_BORDER
        ws_summary.column_dimensions[get_column_letter(col_idx)].width = w

    summary_row = 6
    for pillar_name, controls in ALL_CONTROLS.items():
        ws_summary.cell(row=summary_row, column=1, value=pillar_name).font = DATA_FONT
        ws_summary.cell(row=summary_row, column=1).border = THIN_BORDER
        ws_summary.cell(row=summary_row, column=2, value=len(controls)).font = DATA_FONT
        ws_summary.cell(row=summary_row, column=2).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=summary_row, column=2).border = THIN_BORDER
        # Placeholder counts
        for c in range(3, 6):
            ws_summary.cell(row=summary_row, column=c, value=0).font = DATA_FONT
            ws_summary.cell(row=summary_row, column=c).alignment = Alignment(horizontal="center")
            ws_summary.cell(row=summary_row, column=c).border = THIN_BORDER
        summary_row += 1

    # Total row
    ws_summary.cell(row=summary_row, column=1, value="TOTAL").font = DATA_FONT_BOLD
    ws_summary.cell(row=summary_row, column=1).border = THIN_BORDER
    total = sum(len(c) for c in ALL_CONTROLS.values())
    ws_summary.cell(row=summary_row, column=2, value=total).font = DATA_FONT_BOLD
    ws_summary.cell(row=summary_row, column=2).alignment = Alignment(horizontal="center")
    ws_summary.cell(row=summary_row, column=2).border = THIN_BORDER
    for c in range(3, 6):
        ws_summary.cell(row=summary_row, column=c, value=0).font = DATA_FONT_BOLD
        ws_summary.cell(row=summary_row, column=c).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=summary_row, column=c).border = THIN_BORDER

    ws_summary.merge_cells(f"A{summary_row + 2}:E{summary_row + 2}")
    ws_summary[f"A{summary_row + 2}"].value = "Update the counts above as you complete controls on each pillar sheet."
    ws_summary[f"A{summary_row + 2}"].font = Font(name="Aptos", italic=True, size=10, color=MED_GREY)

    add_footer(ws_summary, summary_row + 4)
    set_print_settings(ws_summary, "Governance Maturity Dashboard \u2014 Summary")

    # Per-pillar sheets
    for pillar_name, controls in ALL_CONTROLS.items():
        # Short sheet name (Excel 31-char limit)
        short = pillar_name.replace("\u2014 ", "").replace(" ", "")[:31]
        ws = wb.create_sheet(title=short)

        # Title
        ws.merge_cells("A1:E1")
        ws["A1"].value = "Governance Maturity Dashboard"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 36

        ws.merge_cells("A2:E2")
        pillar_subtitle = pillar_name + "  |  " + str(len(controls)) + " Controls"
        ws["A2"].value = pillar_subtitle
        ws["A2"].font = SUBTITLE_FONT
        ws.row_dimensions[2].height = 22

        ws.row_dimensions[3].height = 6

        # Header
        header_row = 4
        style_header_row(ws, header_row)

        # Data
        data_start = header_row + 1
        row = data_start
        for i, (cid, cname) in enumerate(controls):
            write_data_row(ws, row, cid, cname, is_even=(i % 2 == 0))
            row += 1

        add_status_validation(ws, data_start, row - 1)
        freeze_and_filter(ws, header_row)
        set_print_settings(ws, "Governance Maturity Dashboard \u2014 " + pillar_name)
        add_footer(ws, row + 1)

    path = os.path.join("docs", "downloads", "governance-maturity-dashboard.xlsx")
    wb.save(path)
    total = sum(len(c) for c in ALL_CONTROLS.values())
    sheets = len(ALL_CONTROLS) + 1  # +1 for Summary
    print(f"  Created: governance-maturity-dashboard.xlsx ({total} controls, {sheets} sheets)")


# ── Main ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("Generating FSI-AgentGov Excel templates...\n")

    for fname, config in ROLE_TEMPLATES.items():
        create_role_checklist(fname, config)
    create_dashboard()

    # Verify
    import zipfile
    print("\nVerification:")
    for f in sorted(os.listdir("docs/downloads")):
        if f.endswith(".xlsx"):
            path = os.path.join("docs/downloads", f)
            is_zip = zipfile.is_zipfile(path)
            size = os.path.getsize(path)
            status = "OK (OOXML)" if is_zip else "FAIL (not OOXML)"
            print(f"  {f}: {status}, {size:,} bytes")

    print("\nDone.")
